"""
data_manager.py
Lee y procesa los datos del Excel de la liga.
Diseñado para migrar fácilmente a una base de datos en el futuro.
"""
import json
import openpyxl
from pathlib import Path

BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data"
PLAYER_PROFILES_PATH = DATA_DIR / "player_profiles.json"

# Mapa: número de jornada → nombre de hoja en el Excel
SHEET_MAP = {
    1: "Jornada 1",
    2: "Jornada 2 (1)",
    3: "Jornada 3",
    4: "Jornada 4",
    5: "Jornada 5",
}

# Estructura de canchas en cada hoja de jornada
# header_row: fila donde está el encabezado de la cancha
# player_rows: filas de los 4 jugadores
# score_row: fila donde están los marcadores numéricos
CANCHAS_LAYOUT = [
    {"numero": 1, "player_rows": [4, 5, 6, 7],   "score_row": 6},
    {"numero": 2, "player_rows": [11, 12, 13, 14], "score_row": 13},
    {"numero": 3, "player_rows": [18, 19, 20, 21], "score_row": 20},
    {"numero": 4, "player_rows": [25, 26, 27, 28], "score_row": 27},
    {"numero": 5, "player_rows": [32, 33, 34, 35], "score_row": 34},
]


def cargar_config():
    with open(BASE_DIR / "config.json", encoding="utf-8") as f:
        return json.load(f)


def guardar_config(config: dict):
    with open(BASE_DIR / "config.json", "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


def _asegurar_data_dir():
    DATA_DIR.mkdir(exist_ok=True)


def cargar_perfiles_jugadores() -> dict:
    """Devuelve el diccionario de perfiles/fotos por nombre de jugador."""
    _asegurar_data_dir()
    if not PLAYER_PROFILES_PATH.exists():
        return {}

    with open(PLAYER_PROFILES_PATH, encoding="utf-8") as f:
        try:
            data = json.load(f)
            return data if isinstance(data, dict) else {}
        except json.JSONDecodeError:
            return {}


def guardar_perfiles_jugadores(perfiles: dict):
    _asegurar_data_dir()
    with open(PLAYER_PROFILES_PATH, "w", encoding="utf-8") as f:
        json.dump(perfiles, f, ensure_ascii=False, indent=2)


def get_perfil_jugador(nombre: str) -> dict:
    perfiles = cargar_perfiles_jugadores()
    return perfiles.get(nombre, {})


def actualizar_perfil_jugador(nombre: str, cambios: dict):
    perfiles = cargar_perfiles_jugadores()
    perfil = perfiles.get(nombre, {})
    perfil.update(cambios)
    perfiles[nombre] = perfil
    guardar_perfiles_jugadores(perfiles)


def _abrir_excel(data_only=True):
    config = cargar_config()
    ruta = BASE_DIR / config["archivo_excel"]
    return openpyxl.load_workbook(ruta, data_only=data_only)


def cargar_jugadores() -> list[dict]:
    """Devuelve la lista de jugadores inscritos desde Hoja2."""
    wb = _abrir_excel()
    ws = wb["Hoja2"]
    perfiles = cargar_perfiles_jugadores()
    jugadores = []
    for row in ws.iter_rows(min_row=2, max_row=40, values_only=True):
        nombre = row[1] if len(row) > 1 else None
        if not nombre:
            continue
        nombre_limpio = str(nombre).strip()
        perfil = perfiles.get(nombre_limpio, {})
        jugadores.append(
            {
                "cancha_base": row[0],
                "nombre": nombre_limpio,
                "presente": bool(row[4]) if len(row) > 4 and row[4] else False,
                "foto_original": perfil.get("foto_original", ""),
                "foto_sin_fondo": perfil.get("foto_sin_fondo", ""),
            }
        )
    return jugadores


def cargar_jornada(numero: int) -> list[dict] | None:
    """
    Devuelve los resultados de una jornada como lista de canchas.
    Cada cancha tiene: numero, jugadores (nombre, puntos, rank_cancha), sets (tuplas de scores).
    """
    if numero not in SHEET_MAP:
        return None

    wb = _abrir_excel()
    sheet_name = SHEET_MAP[numero]
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    canchas = []

    for layout in CANCHAS_LAYOUT:
        jugadores = []
        for row_num in layout["player_rows"]:
            nombre = ws.cell(row=row_num, column=4).value
            puntos = ws.cell(row=row_num, column=3).value
            rank = ws.cell(row=row_num, column=5).value

            if not nombre or not isinstance(nombre, str):
                continue

            jugadores.append(
                {
                    "nombre": nombre.strip(),
                    "puntos": puntos if isinstance(puntos, (int, float)) else 0,
                    "rank_cancha": int(rank) if isinstance(rank, (int, float)) else 99,
                }
            )

        if not jugadores:
            continue

        # Leer scores de los 3 sets (fila score_row)
        sr = layout["score_row"]
        def _val(col):
            v = ws.cell(row=sr, column=col).value
            return int(v) if isinstance(v, (int, float)) else 0

        sets = [
            (_val(6),  _val(8)),   # Set 1: cols F, H
            (_val(10), _val(12)),  # Set 2: cols J, L
            (_val(14), _val(16)),  # Set 3: cols N, P
        ]

        # Ordenar por rank_cancha (1=mejor)
        jugadores.sort(key=lambda j: j["rank_cancha"])

        canchas.append(
            {
                "numero": layout["numero"],
                "jugadores": jugadores,
                "sets": sets,
            }
        )

    return canchas if canchas else None


def cargar_ranking_general() -> list[dict]:
    """
    Lee el ranking acumulado desde la hoja 'Ranking General'.
    Estructura: Jugador, J1, J2, J3, J4, J5, Total, Lugar
    """
    wb = _abrir_excel()
    if "Ranking General" not in wb.sheetnames:
        return []

    ws = wb["Ranking General"]
    ranking = []

    for row in ws.iter_rows(min_row=2, max_row=50, values_only=True):
        nombre = row[0]
        if not nombre or not isinstance(nombre, str):
            continue

        def _p(v):
            return int(v) if isinstance(v, (int, float)) else 0

        pts_por_jornada = {
            1: _p(row[1]) if len(row) > 1 else 0,
            2: _p(row[2]) if len(row) > 2 else 0,
            3: _p(row[3]) if len(row) > 3 else 0,
            4: _p(row[4]) if len(row) > 4 else 0,
            5: _p(row[5]) if len(row) > 5 else 0,
        }
        total = _p(row[6]) if len(row) > 6 else sum(pts_por_jornada.values())
        lugar = _p(row[7]) if len(row) > 7 else 0

        ranking.append(
            {
                "nombre": nombre.strip(),
                "pts_por_jornada": pts_por_jornada,
                "total": total,
                "lugar": lugar,
                "jornadas_jugadas": sum(1 for v in pts_por_jornada.values() if v != 0),
            }
        )

    ranking.sort(key=lambda r: (-r["total"], r["nombre"]))
    for i, r in enumerate(ranking, 1):
        r["posicion"] = i

    return ranking


def get_jornadas_disponibles() -> list[int]:
    """Devuelve los números de jornada que tienen datos en el Excel."""
    wb = _abrir_excel()
    return [n for n, sheet in SHEET_MAP.items() if sheet in wb.sheetnames]
